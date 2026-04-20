"""
reports/excel_report.py - Генератори ҳисоботи Excel
"""
import logging
from datetime import date, timedelta
from pathlib import Path
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import settings

logger = logging.getLogger(__name__)

# --- Стилҳо ---
COLOR_HEADER_BG = "1F3864"
COLOR_HEADER_FG = "FFFFFF"
COLOR_LATE = "FFD700"
COLOR_ABSENT = "FF6B6B"
COLOR_PRESENT = "90EE90"
COLOR_EARLY = "FFA500"
COLOR_ALT_ROW = "F0F4F8"
COLOR_TOTAL_BG = "2E75B6"
COLOR_TOTAL_FG = "FFFFFF"

STATUS_LABELS = {
    "present": "✅ Present",
    "absent": "❌ Absent",
    "late": "⏰ Late",
    "early_leave": "🚪 Early Departure",
    "late_and_early": "⚠️ Late & Early",
    None: "—",
}

WORKDAY_TJ = {
    "Mon": "Душанбе",
    "Tue": "Сешанбе",
    "Wed": "Чоршанбе",
    "Thu": "Панҷшанбе",
    "Fri": "Ҷумъа",
    "Sat": "Шанбе",
    "Sun": "Якшанбе",
}


def _header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)


def _header_font():
    return Font(bold=True, color=COLOR_HEADER_FG, name="Calibri", size=11)


def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_header(ws, row: int, columns: List[str]):
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = _header_fill()
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = _border()


def _row_fill(row_num: int, status: Optional[str] = None) -> PatternFill:
    if status == "absent":
        return PatternFill("solid", fgColor=COLOR_ABSENT)
    elif status == "late":
        return PatternFill("solid", fgColor=COLOR_LATE)
    elif status == "early_leave":
        return PatternFill("solid", fgColor=COLOR_EARLY)
    elif status == "late_and_early":
        return PatternFill("solid", fgColor=COLOR_LATE)
    elif row_num % 2 == 0:
        return PatternFill("solid", fgColor=COLOR_ALT_ROW)
    return PatternFill("solid", fgColor="FFFFFF")


def _fmt_time(val: Optional[str]) -> str:
    if not val:
        return "—"
    try:
        return val.split(" ")[1][:5]  # HH:MM
    except Exception:
        return str(val)


def generate_daily_report(
    records: List[Dict],
    report_date: date,
    output_dir: Path = None,
) -> Path:
    """Daily attendance Excel report"""
    output_dir = output_dir or settings.REPORTS_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    NCOLS = 9
    merge_to = get_column_letter(NCOLS)

    fill_title  = PatternFill("solid", fgColor="0D3B66")
    fill_total  = PatternFill("solid", fgColor="0D3B66")
    font_title  = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    font_total  = Font(bold=True, size=11, color="FFFFFF", name="Calibri")

    # ── Title ────────────────────────────────────────────
    day_name = report_date.strftime("%A")
    ws.merge_cells(f"A1:{merge_to}1")
    c = ws["A1"]
    c.value     = f"DAILY ATTENDANCE REPORT   |   {report_date.strftime('%d.%m.%Y')}  ({day_name})"
    c.fill      = fill_title
    c.font      = font_title
    c.alignment = _center()
    ws.row_dimensions[1].height = 32

    # ── Column headers ───────────────────────────────────
    COLS = [
        "#", "Full Name", "Position",
        "Arrival", "Departure",
        "Late (min)", "Early Leave (min)", "Missed Time", "Status",
    ]
    _apply_header(ws, 2, COLS)
    ws.row_dimensions[2].height = 20

    # ── Data rows ────────────────────────────────────────
    totals = {"late_min": 0, "early_min": 0, "present": 0, "absent": 0}

    for idx, rec in enumerate(records, 1):
        row    = 2 + idx
        status = rec.get("status")
        late_m = int(rec.get("late_minutes")   or 0)
        early_m= int(rec.get("early_leave_min") or 0)
        missed = late_m + early_m
        fill   = _row_fill(idx, status)

        status_map = {
            "present":        "✅ On Time",
            "late":           f"⏰ Late +{late_m}m",
            "early_leave":    f"🚪 Early -{early_m}m",
            "late_and_early": "⚠️ Late & Early",
            "absent":         "❌ Absent",
        }

        values = [
            idx,
            rec.get("full_name", ""),
            rec.get("position") or "—",
            _fmt_time(rec.get("first_in")),
            _fmt_time(rec.get("last_out")),
            late_m  if late_m  else "—",
            early_m if early_m else "—",
            _fmt_mins(missed),
            status_map.get(status, status or "—"),
        ]
        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = _center()
            cell.font      = Font(name="Calibri", size=10)
        ws.row_dimensions[row].height = 18

        if status == "absent":
            totals["absent"] += 1
        else:
            totals["present"] += 1
        totals["late_min"]  += late_m
        totals["early_min"] += early_m

    # ── Total row ────────────────────────────────────────
    total_row = 2 + len(records) + 1
    h_l, m_l = divmod(totals["late_min"],  60)
    h_e, m_e = divmod(totals["early_min"], 60)
    total_missed = totals["late_min"] + totals["early_min"]
    h_t, m_t = divmod(total_missed, 60)

    ws.merge_cells(f"A{total_row}:{merge_to}{total_row}")
    c = ws.cell(total_row, 1,
        f"TOTAL ▸  Present: {totals['present']}  |  Absent: {totals['absent']}  |  "
        f"Total late: {h_l}h {m_l}m  |  Total early leave: {h_e}h {m_e}m  |  "
        f"Total missed: {h_t}h {m_t}m"
    )
    c.fill      = fill_total
    c.font      = font_total
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border    = _border()
    ws.row_dimensions[total_row].height = 24

    # ── Column widths ────────────────────────────────────
    col_widths = [5, 28, 18, 12, 12, 12, 18, 14, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"

    filename = output_dir / f"daily_{report_date.strftime('%Y-%m-%d')}.xlsx"
    wb.save(str(filename))
    logger.info(f"Daily report saved: {filename}")
    return filename


def _fmt_mins(minutes: int) -> str:
    if not minutes:
        return "—"
    h, m = divmod(int(minutes), 60)
    return f"{h}h {m}m" if h else f"{m}m"


def generate_period_report(
    records: List[Dict],
    start_date: date,
    end_date: date,
    report_type: str = "weekly",
    output_dir: Path = None,
) -> Path:
    """Weekly / Monthly report — each employee with all dates + summary"""
    output_dir = output_dir or settings.REPORTS_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    from collections import defaultdict, OrderedDict

    # ── Group records by employee ────────────────────────
    emp_days: Dict[str, Dict] = OrderedDict()
    for rec in sorted(records, key=lambda r: (r.get("full_name", ""), r.get("attendance_date", ""))):
        eid = rec.get("employee_id", "")
        if eid not in emp_days:
            emp_days[eid] = {
                "full_name": rec.get("full_name", ""),
                "position":  rec.get("position") or "—",
                "days": [],
            }
        emp_days[eid]["days"].append(rec)

    # ── Workbook ─────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    label = "WEEKLY" if report_type == "weekly" else "MONTHLY"
    ws.title = f"{label} Report"

    # Styles
    NCOLS = 8
    merge_to = get_column_letter(NCOLS)

    fill_emp_hdr  = PatternFill("solid", fgColor="1F3864")
    fill_col_hdr  = PatternFill("solid", fgColor="2E75B6")
    fill_summary  = PatternFill("solid", fgColor="D6E4F0")
    fill_absent   = PatternFill("solid", fgColor="FFB3B3")
    fill_late     = PatternFill("solid", fgColor="FFF2CC")
    fill_early    = PatternFill("solid", fgColor="FFE0B2")
    fill_late_early = PatternFill("solid", fgColor="FFD700")
    fill_ok       = PatternFill("solid", fgColor="E2EFDA")
    fill_alt      = PatternFill("solid", fgColor="F7FBFF")
    fill_title    = PatternFill("solid", fgColor="0D3B66")
    fill_grand    = PatternFill("solid", fgColor="0D3B66")

    font_title   = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    font_emp     = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    font_col_hdr = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
    font_normal  = Font(size=10, name="Calibri")
    font_summary = Font(bold=True, size=10, color="1F3864", name="Calibri")
    font_grand   = Font(bold=True, size=11, color="FFFFFF", name="Calibri")

    DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    COL_HDR = [
        "Date", "Day", "Arrival", "Departure",
        "Late (min)", "Early Leave (min)", "Missed Time", "Status",
    ]

    cur_row = 1

    # ── Main title ───────────────────────────────────────
    period_str = f"{start_date.strftime('%d.%m.%Y')} – {end_date.strftime('%d.%m.%Y')}"
    ws.merge_cells(f"A{cur_row}:{merge_to}{cur_row}")
    c = ws.cell(cur_row, 1, f"{label} ATTENDANCE REPORT   |   {period_str}")
    c.fill = fill_title; c.font = font_title; c.alignment = _center()
    ws.row_dimensions[cur_row].height = 32
    cur_row += 1

    # ── Grand summary accumulators ───────────────────────
    grand = {"total": 0, "present": 0, "absent": 0,
             "late_days": 0, "early_days": 0,
             "late_min": 0, "early_min": 0}

    # ── Per employee ─────────────────────────────────────
    for emp_num, (eid, emp) in enumerate(emp_days.items(), 1):
        days      = emp["days"]
        name      = emp["full_name"]
        position  = emp["position"]

        # Employee header
        ws.merge_cells(f"A{cur_row}:{merge_to}{cur_row}")
        c = ws.cell(cur_row, 1, f"  {emp_num}.  {name}   |   {position}")
        c.fill = fill_emp_hdr; c.font = font_emp
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        # Column headers
        for ci, h in enumerate(COL_HDR, 1):
            c = ws.cell(cur_row, ci, h)
            c.fill = fill_col_hdr; c.font = font_col_hdr
            c.alignment = _center(); c.border = _border()
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        # Per-day rows
        emp_late_days = emp_early_days = emp_absent = emp_present = 0
        emp_late_min  = emp_early_min  = 0

        for di, rec in enumerate(days):
            status    = rec.get("status") or "absent"
            late_min  = int(rec.get("late_minutes")   or 0)
            early_min = int(rec.get("early_leave_min") or 0)
            missed    = late_min + early_min

            # Row fill
            if status == "absent":
                fill = fill_absent
            elif status == "late_and_early":
                fill = fill_late_early
            elif status == "late":
                fill = fill_late
            elif status == "early_leave":
                fill = fill_early
            elif di % 2 == 0:
                fill = fill_alt
            else:
                fill = fill_ok

            # Date & day name
            att_date = rec.get("attendance_date", "")
            try:
                d = date.fromisoformat(att_date)
                date_fmt = d.strftime("%d.%m.%Y")
                day_name = DAY_NAMES[d.weekday()]
            except Exception:
                date_fmt = att_date
                day_name = "—"

            # Status label
            status_map = {
                "present":        "✅ On time",
                "late":           f"⏰ Late +{late_min}m",
                "early_leave":    f"🚪 Early -{early_min}m",
                "late_and_early": f"⚠️ Late+Early",
                "absent":         "❌ Absent",
            }
            status_label = status_map.get(status, status)

            row_vals = [
                date_fmt,
                day_name,
                _fmt_time(rec.get("first_in")),
                _fmt_time(rec.get("last_out")),
                late_min  if late_min  else "—",
                early_min if early_min else "—",
                _fmt_mins(missed),
                status_label,
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(cur_row, ci, val)
                c.fill = fill; c.border = _border()
                c.alignment = _center(); c.font = font_normal
            ws.row_dimensions[cur_row].height = 18
            cur_row += 1

            # Accumulate
            if status == "absent":
                emp_absent += 1
            else:
                emp_present += 1
            if "late"  in status: emp_late_days  += 1
            if "early" in status: emp_early_days += 1
            emp_late_min  += late_min
            emp_early_min += early_min

        # Employee summary row
        total_missed_min = emp_late_min + emp_early_min
        h_miss, m_miss = divmod(total_missed_min, 60)
        missed_str = f"{h_miss}h {m_miss}m" if total_missed_min else "0"

        ws.merge_cells(f"A{cur_row}:{merge_to}{cur_row}")
        summary_text = (
            f"SUMMARY ▸  Present: {emp_present}  |  Absent: {emp_absent}  |  "
            f"Late days: {emp_late_days}  |  Early-leave days: {emp_early_days}  |  "
            f"Total missed: {missed_str}"
        )
        c = ws.cell(cur_row, 1, summary_text)
        c.fill = fill_summary; c.font = font_summary
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border()
        ws.row_dimensions[cur_row].height = 20
        cur_row += 2   # blank line between employees

        # Grand totals
        grand["total"]      += len(days)
        grand["present"]    += emp_present
        grand["absent"]     += emp_absent
        grand["late_days"]  += emp_late_days
        grand["early_days"] += emp_early_days
        grand["late_min"]   += emp_late_min
        grand["early_min"]  += emp_early_min

    # ── Grand total row ──────────────────────────────────
    total_missed = grand["late_min"] + grand["early_min"]
    h_t, m_t = divmod(total_missed, 60)
    ws.merge_cells(f"A{cur_row}:{merge_to}{cur_row}")
    grand_text = (
        f"GRAND TOTAL ▸  All days: {grand['total']}  |  "
        f"Present: {grand['present']}  |  Absent: {grand['absent']}  |  "
        f"Late days: {grand['late_days']}  |  Early-leave days: {grand['early_days']}  |  "
        f"Total missed: {h_t}h {m_t}m"
    )
    c = ws.cell(cur_row, 1, grand_text)
    c.fill = fill_grand; c.font = font_grand
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border()
    ws.row_dimensions[cur_row].height = 26

    # ── Column widths ────────────────────────────────────
    col_widths = [14, 8, 12, 14, 13, 18, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"

    filename = output_dir / f"{report_type}_{start_date.strftime('%Y-%m-%d')}_{end_date.strftime('%Y-%m-%d')}.xlsx"
    wb.save(str(filename))
    logger.info(f"Period report saved: {filename}")
    return filename


def get_weekly_dates() -> tuple:
    today = date.today()
    start = today - timedelta(days=today.weekday())
    end = today
    return start, end


def get_monthly_dates() -> tuple:
    today = date.today()
    start = today.replace(day=1)
    return start, today
